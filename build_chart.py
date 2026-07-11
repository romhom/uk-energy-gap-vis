"""
UK Gas Capacity Gap & Offshore Wind / Curtailment Cost Chart
--------------------------------------------------------------
Builds a 3-panel figure from the companion workbook
'UK_Electricity_Generation_and_Capacity_Factor.xlsx':

  Panel 1: UK electricity generation mix, 2000-2025 (stacked area)
  Panel 2: Growth in unused gas capacity since 2000 (symmetric "gap" band)
  Panel 3: Offshore wind generation vs. grid constraint/curtailment cost,
           annual totals 2017-2025 (scatter, coloured by year)

Requires: pandas, numpy, matplotlib, openpyxl
Uses the Poppins font bundled in the local fonts/ folder, so no
system-wide font installation is required.

Run this file directly to build the combined 3-panel chart
(gas_gap_wind_curtailment_chart.png). See build_carousel.py for the
single-panel LinkedIn-carousel version, which reuses the data loading
and panel-drawing functions defined here.
"""

import os
import pandas as pd
import numpy as np
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from matplotlib import font_manager
from matplotlib.path import Path
from matplotlib.lines import Line2D

WORKBOOK = "UK_Electricity_Generation_and_Capacity_Factor.xlsx"
OUTPUT_PNG = "gas_gap_wind_curtailment_chart.png"

# ---------------------------------------------------------------
# Fonts
# ---------------------------------------------------------------
FONTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts")
for f in os.listdir(FONTS_DIR):
    if f.lower().endswith(".ttf"):
        font_manager.fontManager.addfont(os.path.join(FONTS_DIR, f))
plt.rcParams["font.family"] = "Poppins"


# ---------------------------------------------------------------
# Custom hand-drawn icon paths (unit-scale, centred at origin)
# used in the generation-mix legend since no reliable wind-turbine /
# flame / oil-drop / wave glyphs exist across common fonts
# ---------------------------------------------------------------
def turbine_path():
    verts = [(0, -0.9), (0, 0.6)]
    codes = [Path.MOVETO, Path.LINETO]
    for ang in [100, 220, 340]:
        rad = np.radians(ang)
        x2, y2 = 0.55 * np.cos(rad), 0.6 + 0.55 * np.sin(rad)
        verts += [(0, 0.6), (x2, y2)]
        codes += [Path.MOVETO, Path.LINETO]
    return Path(verts, codes)


def flame_path():
    verts = [
        (0, -0.9),
        (0.35, -0.3),
        (0.15, 0.1),
        (0.4, 0.5),
        (0, 0.9),
        (-0.4, 0.5),
        (-0.15, 0.1),
        (-0.35, -0.3),
        (0, -0.9),
    ]
    codes = [Path.MOVETO] + [Path.LINETO] * (len(verts) - 1)
    return Path(verts, codes)


def droplet_path():
    verts = [(0, 0.9), (0.5, -0.1), (0.3, -0.7), (-0.3, -0.7), (-0.5, -0.1), (0, 0.9)]
    codes = [Path.MOVETO] + [Path.LINETO] * (len(verts) - 1)
    return Path(verts, codes)


def wave_path():
    xs = np.linspace(-0.9, 0.9, 40)
    ys = 0.35 * np.sin(xs * 3.6)
    verts = list(zip(xs, ys))
    codes = [Path.MOVETO] + [Path.LINETO] * (len(verts) - 1)
    return Path(verts, codes)


ICONS = {
    "Coal": dict(marker="o", text=None),
    "Oil": dict(marker=droplet_path(), text=None),
    "Gas": dict(marker=flame_path(), text=None),
    "Nuclear": dict(marker=None, text="☢"),  # radioactive sign
    "Bioenergy": dict(marker=None, text="☘"),  # shamrock
    "Solar": dict(marker=None, text="☀"),  # black sun with rays
    "Wind (on+offshore)": dict(marker=turbine_path(), text=None),
    "Other/hydro/storage": dict(marker=wave_path(), text=None),
}

STACK_COLORS = {
    "Coal": "#A8A29E",
    "Oil": "#D6A99D",
    "Gas": "#F0A784",
    "Nuclear": "#C4B5E0",
    "Bioenergy": "#A8C4A2",
    "Other/hydro/storage": "#9FD8D3",
    "Solar": "#F2CB6A",
    "Wind (on+offshore)": "#8EC5E8",
}
STACK_ORDER = [
    "Coal",
    "Oil",
    "Gas",
    "Nuclear",
    "Bioenergy",
    "Other/hydro/storage",
    "Solar",
    "Wind (on+offshore)",
]

# ---------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------
FS_LABEL, FS_TICK, FS_LEGEND, FS_ANNOT, FS_ERA = 19, 15.5, 16.5, 14, 17
LW_THICK, LW_MED = 2.8, 2.3
XLIM = (
    2000,
    2025.6,
)  # both top panels stop consistently at 2025 (small margin for the 2025 guide line)

PASTEL_WARM = mcolors.LinearSegmentedColormap.from_list(
    "pastel_warm", ["#FDEBC8", "#F5C88F", "#E0A972", "#C98A5E"]
)


def load_sheet(wb_path, sheet_name, header_row=4, max_row=None):
    """Load a data sheet from the workbook as a plain DataFrame."""
    from openpyxl import load_workbook

    wb = load_workbook(wb_path, data_only=True)
    ws = wb[sheet_name]
    headers = [c.value for c in ws[header_row]]
    rows = list(ws.iter_rows(min_row=header_row + 1, max_row=max_row, values_only=True))
    df = pd.DataFrame(rows, columns=headers)
    df = df[df["Year"].notna()]
    return df


def load_data(wb_path=WORKBOOK):
    """Load and prepare all series needed by the three panels.

    Returns a dict with:
      gen_annual  - DataFrame, annual generation by fuel (STACK_ORDER cols), 2000-2025
      gas         - DataFrame, quarterly gas capacity-gap series, 2000-2025
      scatter_df  - DataFrame, annual offshore wind TWh vs. curtailment cost, 2017-2025
      norm        - matplotlib Normalize instance for scatter_df['year']
    """
    gen = load_sheet(wb_path, "Generation (TWh)", max_row=109)
    cap = load_sheet(wb_path, "Capacity (MW)", max_row=109)
    cf = load_sheet(wb_path, "Capacity Factor (%)", max_row=109)
    curt = load_sheet(wb_path, "Curtailment Costs (GBP)", max_row=40)

    for df in (gen, cap, cf, curt):
        df["period"] = df["Year"] + (df["Quarter"] - 1) / 4

    numeric_cols = [
        "Coal",
        "Oil",
        "Gas",
        "Nuclear",
        "Hydro (natural flow)",
        "Onshore wind",
        "Offshore wind",
        "Shoreline wave / tidal",
        "Solar",
        "Bioenergy",
        "Other fuels",
        "Pumped storage",
        "Battery storage",
    ]
    for c in numeric_cols:
        gen[c] = pd.to_numeric(gen[c], errors="coerce").fillna(0)

    gen["Wind (on+offshore)"] = gen["Onshore wind"] + gen["Offshore wind"]
    gen["Other/hydro/storage"] = (
        gen["Hydro (natural flow)"]
        + gen["Shoreline wave / tidal"]
        + gen["Other fuels"]
        + gen["Pumped storage"]
        + gen["Battery storage"]
    )

    gen_annual_full = gen.groupby(gen["period"].apply(int))[STACK_ORDER].sum()
    gen_annual = gen_annual_full[
        gen_annual_full.index < 2026
    ]  # 2000-2025 only (2026 is partial-year in the source)

    # ---- Gas capacity gap (relative to 2000 baseline), smoothed ----
    gas = gen[["period", "Gas"]].rename(columns={"Gas": "gen_twh"})
    gas["hours"] = cf["Hours in quarter"]
    gas["cap_gw"] = cap["Gas"] / 1000
    gas["avg_output_gw"] = gas["gen_twh"] * 1000 / gas["hours"]
    gas["gap"] = gas["cap_gw"] - gas["avg_output_gw"]
    gas["gap_smooth"] = gas["gap"].rolling(window=4, center=True, min_periods=1).mean()
    baseline = gas.loc[gas["period"] < 2001, "gap_smooth"].mean()
    gas["gap_relative"] = (gas["gap_smooth"] - baseline).clip(lower=0)
    gas["half_gap_relative"] = gas["gap_relative"] / 2
    gas = gas[gas["period"] < 2026]

    # ---- Offshore wind generation vs. constraint cost, annual, 2017-2025 ----
    gen_annual_wind = gen.groupby(gen["period"].apply(int))["Offshore wind"].sum()
    curt_annual = curt.groupby(curt["Year"])["Total cost (GBP)"].sum()
    years = sorted(set(gen_annual_wind.index) & set(curt_annual.index))
    years = [y for y in years if y < 2026]
    scatter_df = pd.DataFrame(
        {
            "year": years,
            "wind_twh": [gen_annual_wind[y] for y in years],
            "cost_m": [curt_annual[y] / 1e6 for y in years],
        }
    )
    norm = mcolors.Normalize(
        vmin=scatter_df["year"].min(), vmax=scatter_df["year"].max()
    )

    return dict(gen_annual=gen_annual, gas=gas, scatter_df=scatter_df, norm=norm)


def style_spines(ax):
    for spine in ax.spines.values():
        spine.set_visible(True)
        spine.set_color("#9CA3AF")
        spine.set_linewidth(1.8)


def draw_panel0(ax, data, show_legend=True, show_xlabel=False):
    """Panel 1: generation mix, stacked area."""
    gen_annual = data["gen_annual"]
    ax.stackplot(
        gen_annual.index,
        [gen_annual[c] for c in STACK_ORDER],
        colors=[STACK_COLORS[c] for c in STACK_ORDER],
        alpha=0.95,
    )
    ax.set_ylabel("Generation (TWh)", fontsize=FS_LABEL)
    if show_xlabel:
        ax.set_xlabel("Year", fontsize=FS_LABEL)
    style_spines(ax)
    ax.grid(alpha=0.15)
    xlim_tmp=(XLIM[0], XLIM[1]-0.6)
    print(xlim_tmp)
    ax.set_xlim(*xlim_tmp)
    ax.tick_params(labelbottom=not show_xlabel or True, labelsize=FS_TICK)
    if not show_xlabel:
        ax.tick_params(labelbottom=False)

    if not show_legend:
        return None

    legend_handles = []
    for c in STACK_ORDER:
        ic = ICONS[c]
        if ic["text"]:
            h = Line2D(
                [0],
                [0],
                marker="$" + ic["text"] + "$",
                color="none",
                markerfacecolor=STACK_COLORS[c],
                markeredgecolor=STACK_COLORS[c],
                markersize=24,
                label=c,
            )
        else:
            h = Line2D(
                [0],
                [0],
                marker=ic["marker"],
                color="none",
                markerfacecolor=STACK_COLORS[c],
                markeredgecolor="#6B7280",
                markeredgewidth=1.5,
                markersize=19,
                label=c,
            )
        legend_handles.append(h)

    leg = ax.legend(
        handles=legend_handles,
        loc="lower center",
        bbox_to_anchor=(0.5, 1.0),
        frameon=True,
        fontsize=FS_LEGEND,
        ncol=4,
        handletextpad=0.5,
        columnspacing=1.4,
        borderaxespad=0.8,
    )
    leg.get_frame().set_edgecolor("#9CA3AF")
    leg.get_frame().set_linewidth(1.6)
    leg.get_frame().set_boxstyle("round,pad=0.5")
    return leg


def draw_panel1(ax, data):
    """Panel 2: gas capacity gap, with historical eras shaded."""
    gas = data["gas"]
    scatter_df = data["scatter_df"]

    ax.axvspan(2000, 2010, color="#EDEEF0", alpha=0.6, zorder=0)
    ax.axvspan(2010, 2019, color="#E1F1E5", alpha=0.6, zorder=0)
    ax.axvspan(2019, 2025.6, color="#E4EDF9", alpha=0.6, zorder=0)
    ax.fill_between(
        gas["period"],
        -gas["half_gap_relative"],
        gas["half_gap_relative"],
        color="#F0A784",
        alpha=0.55,
        zorder=2,
    )
    ax.plot(
        gas["period"],
        gas["half_gap_relative"],
        color="#D98B60",
        linewidth=LW_THICK,
        zorder=2,
    )
    ax.plot(
        gas["period"],
        -gas["half_gap_relative"],
        color="#D98B60",
        linewidth=LW_THICK,
        zorder=2,
    )
    ax.axhline(0, color="#6B7280", linewidth=LW_THICK, zorder=2)
    for y in scatter_df["year"]:
        color = PASTEL_WARM(data["norm"](y))
        ax.axvline(y, color=color, linewidth=LW_MED, alpha=0.9, zorder=1)
    ax.set_ylabel("Growth in unused gas\ncapacity since 2000 (GW)", fontsize=FS_LABEL)
    ax.set_xlabel("Year", fontsize=FS_LABEL)
    style_spines(ax)
    ax.grid(alpha=0.15, zorder=0)
    ax.set_xlim(*XLIM)
    ax.tick_params(labelsize=FS_TICK)
    ymax = ax.get_ylim()[1]
    for y in scatter_df["year"]:
        ax.text(
            y,
            ymax * 0.92,
            str(int(y)),
            fontsize=FS_ANNOT - 1.5,
            color="#B06A3F",
            ha="center",
            rotation=90,
            va="top",
            fontweight="bold",
            zorder=3,
        )
    ymin = ax.get_ylim()[0]
    era_y = ymin + (ymax - ymin) * 0.06
    ax.text(
        2005,
        era_y,
        "Pre-wind CCGT buildout",
        fontsize=FS_ERA,
        ha="center",
        color="#6B7280",
        style="italic",
    )
    ax.text(
        2014.5,
        era_y,
        "Demand decline +\nearly wind growth",
        fontsize=FS_ERA,
        ha="center",
        color="#4B8065",
        style="italic",
    )
    ax.text(
        2022.3,
        era_y,
        "Wind at scale",
        fontsize=FS_ERA,
        ha="center",
        color="#4A6FA5",
        style="italic",
    )


def draw_panel2(ax, data, fig=None):
    """Panel 3: offshore wind generation vs. constraint cost, scatter."""
    scatter_df = data["scatter_df"]
    norm = data["norm"]

    style_spines(ax)
    sc = ax.scatter(
        scatter_df["wind_twh"],
        scatter_df["cost_m"],
        c=scatter_df["year"],
        cmap=PASTEL_WARM,
        norm=norm,
        s=300,
        edgecolor="#6B7280",
        linewidth=1.4,
        zorder=3,
    )
    for _, row in scatter_df.iterrows():
        yr = int(row["year"])
        xytext = (-55, -6) if yr == 2025 else (9, 7)
        ax.annotate(
            str(yr),
            (row["wind_twh"], row["cost_m"]),
            fontsize=FS_ANNOT + 1.5,
            xytext=xytext,
            textcoords="offset points",
            color="#4B5563",
            fontweight="bold",
        )

    # directional guide arrow (thicker shaft + enlarged arrowhead via mutation_scale)
    ax.annotate(
        "",
        xy=(48, 1950),
        xytext=(16, 420),
        arrowprops=dict(
            arrowstyle="-|>",
            color="#9CA3AF",
            lw=3.6,
            alpha=0.75,
            connectionstyle="arc3,rad=0.15",
            mutation_scale=32,
        ),
    )
    ax.text(
        20,
        750,
        "More wind\nHigher curtailment cost",
        fontsize=FS_ANNOT,
        color="#6B7280",
        style="italic",
        ha="left",
        va="center",
    )

    ax.set_xlabel("Offshore wind generation (TWh, annual)", fontsize=FS_LABEL)
    ax.set_ylabel("Constraint/curtailment cost (£m, annual)", fontsize=FS_LABEL)
    ax.grid(alpha=0.15)
    ax.tick_params(labelsize=FS_TICK)

    if fig is not None:
        # horizontal colorbar inset, top-left, ~35% of panel width
        cbar_ax = ax.inset_axes([0.05, 0.83, 0.35, 0.035])
        cbar = fig.colorbar(sc, cax=cbar_ax, orientation="horizontal")
        cbar.ax.set_title("Year", fontsize=FS_LABEL - 3, pad=4)
        cbar.ax.tick_params(labelsize=FS_TICK - 3)
        cbar.outline.set_edgecolor("#9CA3AF")
        cbar.outline.set_linewidth(1.2)


def main():
    data = load_data()

    # ---------------------------------------------------------------
    # Figure
    # ---------------------------------------------------------------
    fig = plt.figure(figsize=(17, 19.6))
    gs = fig.add_gridspec(3, 1, height_ratios=[0.85, 1, 1.25], hspace=0.30)
    ax0 = fig.add_subplot(gs[0])
    ax1 = fig.add_subplot(gs[1])
    ax2 = fig.add_subplot(gs[2])

    draw_panel0(ax0, data)
    draw_panel1(ax1, data)
    draw_panel2(ax2, data, fig=fig)

    # ---- align panel widths/positions: panels 0 & 1 flush together, panel 2 nudged up ----
    fig.canvas.draw()
    pos0 = ax0.get_position()
    pos1 = ax1.get_position()
    pos2 = ax2.get_position()
    common_x0 = pos1.x0
    common_width = pos1.width
    ax0.set_position([common_x0, pos1.y1, common_width, pos0.height])
    ax2.set_position([common_x0, pos2.y0 + 0.008, common_width, pos2.height])

    plt.savefig(OUTPUT_PNG, dpi=150, bbox_inches="tight", pad_inches=0.35)
    print(f"Saved {OUTPUT_PNG}")


if __name__ == "__main__":
    main()
